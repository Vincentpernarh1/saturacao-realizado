import math

import pandas as pd
from tkinter import *
from tkinter import ttk
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill, Font, Alignment
from math import ceil
import re
from PIL import Image, ImageTk
import traceback
import os
import sys
import glob
import numpy as np
import warnings 
import contextlib
import unicodedata

# Suppress xlrd / Excel warnings
warnings.simplefilter("ignore")

# Optionally suppress all print output from the engine
with contextlib.redirect_stdout(None), contextlib.redirect_stderr(None):
    df = pd.read_excel("Template.xlsx", engine="openpyxl")
warnings.filterwarnings(
    "ignore",
    category=UserWarning,
    message=".*OLE2.*"
)
warnings.filterwarnings(
    "ignore",
    category=UserWarning,
    message=".*CODEPAGE.*"
)
warnings.filterwarnings(
    "ignore",
    message="^WARNING .*" # Hides the file size warnings which don't have a category
)



# Always resolve paths relative to the exe's location (or source root), never os.getcwd()
if getattr(sys, 'frozen', False):
    caminho_base = os.path.dirname(sys.executable)
else:
    caminho_base = os.path.dirname(os.path.abspath(__file__))

# Helper function to find latest file matching pattern
def get_latest_file(pattern, fallback=None):
    """Find the most recent file matching the pattern, or return fallback file if none found.
    
    Args:
        pattern: Glob pattern to match files (e.g., "BD/BD_CADASTRO_PN_*.xlsx")
        fallback: Fallback filename if no dated file is found (e.g., "BD/BD_CADASTRO_PN.xlsx")
        
    Returns:
        str: Path to the latest file, or fallback, or None
    """
    files = glob.glob(pattern)
    if files:
        # Sort by modification time, newest first
        files.sort(key=os.path.getmtime, reverse=True)
        return files[0]
    elif fallback and os.path.exists(fallback):
        return fallback
    return None

# ==================== DEBUG CONFIGURATION ====================
# Add supplier codes here to debug mapping issues
# You can use either COD IMS or COD FORNECEDOR (or both if a supplier has both codes)
# Note: A single supplier can have BOTH COD IMS and COD FORNECEDOR
# Example: Same supplier with COD IMS='33611' and COD FORNECEDOR='800030798'
DEBUG_SUPPLIERS = []  # Add your supplier codes here to debug (e.g., ['33611', '800030798'])

# Add AGRUPAMENTO codes to filter debug output (optional, leave empty to see all)
# Example: DEBUG_AGRUPAMENTO = ['30956207']
DEBUG_AGRUPAMENTO = []  # Add AGRUPAMENTO codes to filter debug output

# --- ADDED DEBUG CONFIGURATION FOR PNs ---
DEBUG_PNS = []  # Add PN (DESENHO) codes to debug missing PNs/matches (e.g., ['12345678'])
# -----------------------------------------

# Global list to store debug information
debug_info = []

def add_debug_info(supplier_code, stage, message, details=None, agrupamento=None):
    """Add debug information for tracked suppliers."""
    if not DEBUG_SUPPLIERS:
        return
    
    # Check if this supplier should be debugged
    supplier_str = str(supplier_code).strip()
    should_debug = any(str(debug_code).strip() in supplier_str or supplier_str in str(debug_code).strip() 
                       for debug_code in DEBUG_SUPPLIERS)
    
    # Also check AGRUPAMENTO filter if specified
    if should_debug and DEBUG_AGRUPAMENTO and agrupamento:
        agrupamento_str = str(agrupamento).strip()
        should_debug = any(str(ag).strip() in agrupamento_str or agrupamento_str in str(ag).strip()
                          for ag in DEBUG_AGRUPAMENTO)
    
    if should_debug:
        debug_entry = {
            'supplier': supplier_str,
            'stage': stage,
            'message': message,
            'details': details or {}
        }
        debug_info.append(debug_entry)
        
        # Print for immediate visibility
        print(f"\n[DEBUG] Supplier: {supplier_str} | Stage: {stage}")
        print(f"        {message}")
        if details:
            for key, value in details.items():
                print(f"        {key}: {value}")

def print_debug_summary():
    """Print a summary of all debug information collected."""
    if not debug_info:
        return
    
    print("\n" + "="*80)
    print("DEBUG SUMMARY FOR TRACKED SUPPLIERS")
    print("="*80)
    
    # Group by supplier
    suppliers = {}
    for entry in debug_info:
        sup = entry['supplier']
        if sup not in suppliers:
            suppliers[sup] = []
        suppliers[sup].append(entry)
    
    for supplier, entries in suppliers.items():
        print(f"\n>>> SUPPLIER: {supplier}")
        print("-" * 80)
        for entry in entries:
            print(f"  [{entry['stage']}] {entry['message']}")
            if entry['details']:
                for key, value in entry['details'].items():
                    print(f"    - {key}: {value}")
        print()
    
    print("="*80 + "\n")

def clear_debug_info():
    """Clear debug information."""
    global debug_info
    debug_info = []

# ==================== END DEBUG CONFIGURATION ====================

# Lista global para coletar erros e avisos para mostrar ao usuário
erros_processamento = []

def adicionar_erro(mensagem, tipo="ERRO"):
    """Adiciona uma mensagem de erro ou aviso à lista global sem duplicatas."""
    msg = f"[{tipo}] {mensagem}"
    if msg not in erros_processamento:
        erros_processamento.append(msg)
        print(msg)

def limpar_erros():
    """Limpa a lista de erros"""
    global erros_processamento
    erros_processamento = []

def obter_erros():
    """Retorna a lista de erros acumulados"""
    return erros_processamento.copy()


def _campo_tem_codigo(campo, alvo):
    if pd.isna(campo):
        return False
    alvo_str = str(alvo).strip()
    if not alvo_str:
        return False
    partes = [codigo.strip() for codigo in re.split(r'\s*[,/]\s*', str(campo).strip()) if codigo.strip()]
    return alvo_str in partes


def _normalizar_codigos_campo(campo):
    if pd.isna(campo):
        return []
    return [codigo.strip() for codigo in re.split(r'\s*[,/]\s*', str(campo).strip()) if codigo.strip()]


def _codigo_principal(campo):
    codigos = _normalizar_codigos_campo(campo)
    if not codigos:
        return ''

    codigo = codigos[0]
    if codigo in ('0', '0.0'):
        return ''

    try:
        return str(int(float(codigo)))
    except (ValueError, TypeError):
        return codigo


def _mdr_chave(campo):
    if pd.isna(campo):
        return ''

    texto = str(campo).strip()
    if texto.lower() in ('', 'nan', 'none'):
        return ''

    return texto.upper()


def _chave_fornecedor_mdr(fornecedor, mdr):
    fornecedor_chave = _codigo_principal(fornecedor)
    mdr_chave = _mdr_chave(mdr)

    if not fornecedor_chave or not mdr_chave:
        return ''

    return f"{fornecedor_chave}|{mdr_chave}"


def normalize_sheet_name(desired_name, available_sheets):
    """
    Finds a matching sheet name from available_sheets that matches desired_name,
    ignoring case and accents.
    
    Args:
        desired_name: The sheet name we want (e.g., 'Sábado', 'Geral', 'Domingo')
        available_sheets: List of actual sheet names in the Excel file
    
    Returns:
        The actual sheet name from the file that matches, or None if no match found
    """
    if desired_name is None or not available_sheets:
        return None
    
    def remove_accents(text):
        """Remove accents from text"""
        if text is None:
            return ''
        text = str(text)
        # Normalize to NFD (decomposed form) and filter out combining characters
        nfd = unicodedata.normalize('NFD', text)
        return ''.join(char for char in nfd if unicodedata.category(char) != 'Mn')
    
    # Normalize the desired name (remove accents, convert to upper)
    desired_normalized = remove_accents(desired_name).upper().strip()
    
    # Try to find a match
    for sheet in available_sheets:
        sheet_normalized = remove_accents(str(sheet)).upper().strip()
        if sheet_normalized == desired_normalized:
            return sheet  # Return the actual sheet name from the file
    
    return None  # No match found
    

def Processar_Demandas(pasta_demandas="MVM", sheet_name=None):
    """
    Reads Excel files from MVM folder and returns all columns as-is.
    No FLUXO matching, no filtering - just read and validate basic data quality.
    
    Args:
        pasta_demandas: Folder name containing Excel files (default: "MVM")
        sheet_name: Optional sheet name to read (Geral, Sábado, Domingo, etc.)
    
    Returns:
        DataFrame with all columns from Excel files
    """
    # Define o caminho completo para a pasta de demandas
    caminho_pasta = os.path.join(caminho_base, pasta_demandas)

    # Verifica se a pasta de demandas existe
    if not os.path.isdir(caminho_pasta):
        adicionar_erro(f"Pasta de demandas não encontrada: '{caminho_pasta}'", "ERRO")
        return pd.DataFrame()

    # Lista para armazenar os DataFrames de cada arquivo processado
    lista_dfs = []

    # Percorre todos os arquivos na pasta de demandas
    for nome_arquivo in os.listdir(caminho_pasta):
        caminho_completo_arquivo = os.path.join(caminho_pasta, nome_arquivo)
        nome_arquivo_lower = nome_arquivo.lower()
        
        try:
            # Process Excel files with "saturação" and "mvm" in the name
            if nome_arquivo_lower.endswith((".xls", ".xlsx")) and ("saturação" in nome_arquivo_lower and "mvm" in nome_arquivo_lower) and not nome_arquivo_lower.startswith("~$"):
                
                # Read the Excel file - keep ALL columns
                df_excel = pd.read_excel(caminho_completo_arquivo, sheet_name=sheet_name or 0)

                # Verify minimum required columns exist
                required_cols = ['DESENHO', 'QTDE']
                missing = [c for c in required_cols if c not in df_excel.columns]
                if missing:
                    adicionar_erro(f"Arquivo '{nome_arquivo}': Colunas obrigatórias faltando: {', '.join(missing)}", "ERRO")
                    continue

                # Keep all columns from the Excel file
                df_temp = df_excel.copy()
                
                # Standardize column name: VEICULO -> VEÍCULO if needed
                if 'VEICULO' in df_temp.columns and 'VEÍCULO' not in df_temp.columns:
                    df_temp.rename(columns={'VEICULO': 'VEÍCULO'}, inplace=True)
                
                # Add to list for concatenation
                lista_dfs.append(df_temp)
                
                
        except Exception as e:
            adicionar_erro(f"Erro ao processar arquivo '{nome_arquivo}': {str(e)}", "ERRO")
            continue

    # --- CONSOLIDATE DATA ---
    if not lista_dfs:
        adicionar_erro("Nenhum dado válido foi processado. Verifique os arquivos na pasta MVM.", "ERRO")
        return pd.DataFrame()
    
    # Concatenate all DataFrames
    df_final = pd.concat(lista_dfs, ignore_index=True)
    # df_final = df.groupby(colums['AGRUPAMENTO']).agg({'QTDE': 'sum'}).reset_index()
    
    # Basic data quality: convert DESENHO and QTDE to numeric
    if 'DESENHO' in df_final.columns:
        df_final['DESENHO'] = pd.to_numeric(df_final['DESENHO'], errors='coerce')
    
    if 'QTDE' in df_final.columns:
        df_final['QTDE'] = pd.to_numeric(df_final['QTDE'], errors='coerce')
    
    # Remove rows with invalid DESENHO or QTDE
    df_final.dropna(subset=["DESENHO", "QTDE"], inplace=True)
    
    # Remove rows where QTDE is zero or negative
    df_final = df_final[df_final['QTDE'] > 0]
    
    # Convert DESENHO and QTDE to int
    if 'DESENHO' in df_final.columns and df_final['DESENHO'].notna().all():
        df_final['DESENHO'] = df_final['DESENHO'].astype(int)
    
    if 'QTDE' in df_final.columns and df_final['QTDE'].notna().all():
        df_final['QTDE'] = df_final['QTDE'].astype(int)
    
    # Clean up string columns: remove .0 suffix from numeric codes stored as strings
    string_code_cols = ['COD FORNECEDOR', 'COD IMS']
    for col in string_code_cols:
        if col in df_final.columns:
            def _clean_code(val):
                if pd.isna(val):
                    return None
                s = str(val).strip()
                if s in ('nan', '', 'None'):
                    return None
                # Keep compound codes like "56589/46051"
                if '/' in s:
                    return s
                # Remove .0 suffix from numeric codes
                try:
                    return str(int(float(s)))
                except (ValueError, TypeError):
                    return s
            df_final[col] = df_final[col].apply(_clean_code)
    
    # --- CONVERT VEHICLE NAMES TO CODES ---
    # Load vehicle mapping from VEÍCULOS.xlsx
    if 'VEÍCULO' in df_final.columns:
        veiculo_map = _load_vehicle_mapping()
        
        if veiculo_map:
            # Map vehicle names to codes
            def map_vehicle_name_to_code(veiculo_name):
                if pd.isna(veiculo_name):
                    return None
                # Try exact match first
                name_str = str(veiculo_name).strip()
                if name_str in veiculo_map:
                    return veiculo_map[name_str]
                # Try uppercase match
                name_upper = name_str.upper()
                if name_upper in veiculo_map:
                    return veiculo_map[name_upper]
                # Try case-insensitive search
                for key, code in veiculo_map.items():
                    if key.upper() == name_upper:
                        return code
                # If not found, try to convert to int (in case it's already a code)
                try:
                    return int(float(name_str))
                except (ValueError, TypeError):
                    adicionar_erro(f"Veículo não encontrado no mapeamento: '{veiculo_name}'", "AVISO")
                    return None
            
            df_final['VEÍCULO'] = df_final['VEÍCULO'].apply(map_vehicle_name_to_code)
            
            # Remove rows where vehicle mapping failed
            rows_before = len(df_final)
            df_final = df_final[df_final['VEÍCULO'].notna()]
            rows_removed = rows_before - len(df_final)
            if rows_removed > 0:
                adicionar_erro(f"{rows_removed} linha(s) removida(s) por falta de mapeamento de veículo", "AVISO")
            
            # Convert to int
            if not df_final.empty:
                df_final['VEÍCULO'] = df_final['VEÍCULO'].astype(int)
        else:
            adicionar_erro("Não foi possível carregar mapeamento de veículos. Tentando converter diretamente.", "AVISO")
            # Try to convert directly to int
            df_final['VEÍCULO'] = pd.to_numeric(df_final['VEÍCULO'], errors='coerce')
            df_final = df_final[df_final['VEÍCULO'].notna()]
            if not df_final.empty:
                df_final['VEÍCULO'] = df_final['VEÍCULO'].astype(int)
    
    return df_final


def _load_vehicle_mapping():
    """
    Load vehicle name to code mapping from VEÍCULOS.xlsx file.
    Returns a dict mapping vehicle names (uppercase) to codes (int).
    """
    # Fallback mapping if file not found
    _FALLBACK_VEICULOS = {
        'BIG SIDER': 6, 'BITREM': 7, 'CARRETA': 4, 'CARRETA LINE HAUL': 14,
        'CARRETA REBAIXADA': 9, 'CTNR 20': 15, 'CTNR 40': 16, 'FIORINO': 11,
        'RODOTREM': 8, 'TRUCK 3M': 3, 'TRUCK 3M ALONGADO': 18, 'TRUCK 3M PLUS': 13,
        'TRUCK ALONGADO': 17, 'TRUCK VIAGEM': 2, 'TRUCK VIAGEM PLUS': 12, 'VAN': 10,
        'VANDERLEA': 5, 'VEÍCULO 3/4': 1, 'TRUCK SIDER': 2
    }
    
    possible_files = [
        os.path.join(caminho_base, "BD", "VEÍCULOS.xlsx"),
        os.path.join(caminho_base, "BD", "VEICULOS.xlsx"),
        os.path.join(caminho_base, "BD", "Veiculos.xlsx"),
        os.path.join(caminho_base, "BD", "VEICULOS.xls")
    ]
    
    for fpath in possible_files:
        if os.path.exists(fpath):
            try:
                df_veh = pd.read_excel(fpath, sheet_name=0, dtype=str)
                # normalize column names (case-insensitive)
                cols = {c.strip().upper(): c for c in df_veh.columns}
                # find code column (prefer "COD VEICULO" or similar)
                code_col = None
                desc_col = None
                for key_upper, orig in cols.items():
                    if "COD" in key_upper and "VEIC" in key_upper:
                        code_col = orig
                    # Look for DESCRICAO column first, then VEICULOS
                    if "DESCRI" in key_upper:
                        desc_col = orig
                    elif desc_col is None and "VEIC" in key_upper and "COD" not in key_upper:
                        desc_col = orig
                # fallback: use first column as code and second as desc
                if code_col is None and len(df_veh.columns) >= 1:
                    code_col = df_veh.columns[0]
                if desc_col is None and len(df_veh.columns) >= 2:
                    desc_col = df_veh.columns[1]
                if code_col is None or desc_col is None:
                    continue
                
                veic_map = {}
                for _, r in df_veh.iterrows():
                    desc = str(r.get(desc_col, "")).strip()
                    code_raw = r.get(code_col, "")
                    # try to convert code to int
                    try:
                        code = int(float(str(code_raw).strip()))
                    except Exception:
                        continue
                    if desc and code:
                        # Store both original and uppercase for case-insensitive lookup
                        veic_map[desc] = code
                        veic_map[desc.upper()] = code
                
                if veic_map:
                    return veic_map
            except Exception as e:
                print(f"[WARN] Could not read vehicles file {fpath}: {e}")
    
    # Return fallback mapping with uppercase keys added
    fallback_with_uppercase = {}
    for k, v in _FALLBACK_VEICULOS.items():
        fallback_with_uppercase[k] = v
        fallback_with_uppercase[k.upper()] = v
    
    return fallback_with_uppercase

# Exemplo de como chamar a função
# df_processado = Processar_Demandas(cod_destino="BR01")
# print(df_processado)




def desenhar_caminhoes(canvas, ocupacao, caminhao_img):
    canvas.delete("all")

    if caminhao_img is None:
        return

    quad_por_caminhao = 35
    quad_linha = 7
    quad_coluna = 5
    quad_largura = 14
    quad_altura = 14
    margem = 10

    total_quads = ceil(ocupacao * quad_por_caminhao / 100)
    max_caminhoes = 3
    num_caminhoes = min((total_quads - 1) // quad_por_caminhao + 1, max_caminhoes)

    for caminhao_idx in range(num_caminhoes):
        # Posição em "grade" 2 acima, 1 abaixo
        if caminhao_idx < 2:
            x_offset = margem + caminhao_idx * 180  # lado a lado
            y_offset = margem
        else:
            x_offset = margem + 90  # centraliza abaixo dos dois
            y_offset = margem + 130

        canvas.create_image(x_offset + 12, y_offset + 17, image=caminhao_img, anchor=NW)

        x_inicial_grade = x_offset + 50
        y_inicial_grade = y_offset + 10

        for i in range(quad_coluna):
            for j in range(quad_linha):
                idx = caminhao_idx * quad_por_caminhao + (quad_coluna - 1 - i) * quad_linha + j
                x1 = x_inicial_grade + j * quad_largura
                y1 = y_inicial_grade + i * quad_altura
                x2 = x1 + quad_largura
                y2 = y1 + quad_altura
                cor = "#0070C0" if idx < total_quads else "#D9D9D9"
                canvas.create_rectangle(x1, y1, x2, y2, fill=cor, outline='black')



def calcular_empilhamento_line_haul(df_saturacao, db_empilhamento):
    empilhamento_rows = []

    base_df = df_saturacao[df_saturacao['EMBALAGEM_BASE'] == 1]
    sobre_df = df_saturacao[df_saturacao['EMBALAGEM_SOBREPOSTA'] == 1]

    for _, base_row in base_df.iterrows():
        for _, sobre_row in sobre_df.iterrows():
            if base_row['COD FORNECEDOR'] == sobre_row['COD FORNECEDOR']:
                fornecedor = base_row['COD FORNECEDOR']
                embal_base = base_row['EMBALAGEM']
                embal_sobre = sobre_row['EMBALAGEM']

                empilhamento_match = db_empilhamento[
                    (db_empilhamento['COD FORNECEDOR'] == fornecedor) &
                    (db_empilhamento['MDR BASE'] == embal_base) &
                    (db_empilhamento['MDR SOBREPOSTA'] == embal_sobre)
                ]

                if empilhamento_match.empty:
                    continue

                capacidade_veiculo = base_row['CAPACIDADE']

                total_base = base_row['TOTAL DE CXS']
                total_sobre = sobre_row['TOTAL DE CXS']

                usadas_base = 0
                usadas_sobre = 0

                # Empilha 1 base com 1 sobreposta (não considera EMPILHAMENTO BASE)
                while total_base >= 1 and total_sobre >= 1:
                    total_base -= 1
                    total_sobre -= 1
                    usadas_base += 1
                    usadas_sobre += 1

                total_empilhado = usadas_base + usadas_sobre
                chave = f"{fornecedor}-{embal_base}-{embal_sobre}"
                
                saturacao = total_empilhado / capacidade_veiculo

                empilhamento_rows.append({
                    'FORNECEDOR': fornecedor,
                    'EMBALAGEM_BASE': embal_base,
                    'EMBALAGEM_SOBREPOSTA': embal_sobre,
                    'CAPACIDADE_VEÍCULO': capacidade_veiculo,
                    'TOTAL_DE_EMBALAGENS_BASE': base_row['TOTAL DE CXS'],
                    'TOTAL_DE_EMBALAGENS_SOBREPOSTA': sobre_row['TOTAL DE CXS'],
                    'TOTAL_DE_EMBALAGENS_BASE_PARA_COMBINAR': usadas_base,
                    'TOTAL_DE_EMBALAGENS_SOBREPOSTA_PARA_COMBINAR': usadas_sobre,
                    'EMBALAGENS_BASE_RESTANTE': total_base,
                    'EMBALAGENS_SOBREPOSTA_RESTANTE': total_sobre,
                    'CHAVE': chave,
                    'TOTAL_EMBALAGENS_EMPILHADAS': total_empilhado,
                    'SATURAÇÃO': saturacao,
                    'EMPILHAMENTO BASE': 1  # fixo, pois é 1:1
                })

    return pd.DataFrame(empilhamento_rows)


def calcular_empilhamento(df_saturacao, db_empilhamento):
    empilhamento_rows = []

    base_df = df_saturacao[df_saturacao['EMBALAGEM_BASE'] == 1]
    sobre_df = df_saturacao[df_saturacao['EMBALAGEM_SOBREPOSTA'] == 1]

    for _, base_row in base_df.iterrows():
        for _, sobre_row in sobre_df.iterrows():
            # Must match on COD FLUXO too — stacking only within the same fluxo
            if base_row['COD FORNECEDOR'] == sobre_row['COD FORNECEDOR'] and base_row['COD FLUXO'] == sobre_row['COD FLUXO']:
                fornecedor = base_row['COD FORNECEDOR']
                cod_fluxo = base_row['COD FLUXO']
                embal_base = base_row['EMBALAGEM']
                embal_sobre = sobre_row['EMBALAGEM']

                empilhamento_match = db_empilhamento[
                    (db_empilhamento['COD FORNECEDOR'] == fornecedor) &
                    (db_empilhamento['MDR BASE'] == embal_base) &
                    (db_empilhamento['MDR SOBREPOSTA'] == embal_sobre)
                ]

                if empilhamento_match.empty:
                    continue

                emp_base = empilhamento_match.iloc[0]['EMPILHAMENTO BASE']
                capacidade_veiculo = base_row['CAPACIDADE']

                total_base = base_row['TOTAL DE CXS']
                total_sobre = sobre_row['TOTAL DE CXS']

                usadas_base = 0
                usadas_sobre = 0

                while total_base >= emp_base and total_sobre >= 1:
                    total_base -= emp_base
                    total_sobre -= 1
                    usadas_base += emp_base
                    usadas_sobre += 1

                total_empilhado = usadas_base + usadas_sobre
                chave = f"{fornecedor}-{embal_base}-{embal_sobre}"
                
                saturacao = total_empilhado / capacidade_veiculo

                empilhamento_rows.append({
                    'FORNECEDOR': fornecedor,
                    'COD FLUXO': cod_fluxo,
                    'EMBALAGEM_BASE': embal_base,
                    'EMBALAGEM_SOBREPOSTA': embal_sobre,
                    'CAPACIDADE_VEÍCULO': capacidade_veiculo,
                    'TOTAL_DE_EMBALAGENS_BASE': base_row['TOTAL DE CXS'],
                    'TOTAL_DE_EMBALAGENS_SOBREPOSTA': sobre_row['TOTAL DE CXS'],
                    'TOTAL_DE_EMBALAGENS_BASE_PARA_COMBINAR': usadas_base,
                    'TOTAL_DE_EMBALAGENS_SOBREPOSTA_PARA_COMBINAR': usadas_sobre,
                    'EMBALAGENS_BASE_RESTANTE': total_base,
                    'EMBALAGENS_SOBREPOSTA_RESTANTE': total_sobre,
                    'CHAVE': chave,
                    'TOTAL_EMBALAGENS_EMPILHADAS': total_empilhado,
                    'SATURAÇÃO': saturacao,
                    'EMPILHAMENTO BASE': emp_base
                })

    return pd.DataFrame(empilhamento_rows)



def completar_informacoes(tree, veiculo, tree_resumo, canvas_caminhoes, caminhao_img, usar_manual=False,caminho_BD = 'BD'):

    def split_key_logic(code):
        """
        Splits a code by '/'. 
        Returns the second element if a split occurs, otherwise returns the original code.
        """
        # Convert to string just in case, then split
        
        parts = str(code).split('/')
        
        if len(parts) > 1:
            # If the split created more than one part, take the second one (index 1)
            return parts[1].strip() 
            
        else:
            # Otherwise, take the original part (index 0)
            return parts[0].strip()
    try:

        # === Clear debug info at the start of processing ===
        clear_debug_info()

        # --- Leitura dos arquivos ---
        template = pd.read_excel('Template.xlsx', dtype={'DESENHO': str})
        
        print(f"Template original: {template.shape[0]} rows")
       
        template = template[template['QTDE'] > 0]
       
        
        # Early validation: check if template is empty after filtering
        if template.empty:
            adicionar_erro("Template está vazio após filtragem. Verifique se os arquivos de demanda possuem dados válidos com COD FLUXO e QTDE > 0.", "ERRO")
            raise ValueError("Template vazio - nenhum dado válido para processar")
        
        # Ensure COD IMS column exists (for backward compatibility with files that don't have it)
        if 'COD IMS' not in template.columns:
            template['COD IMS'] = ""
        
        # Clean up COD FORNECEDOR: always store as string to prevent float upcast ("800006330.0").
        # Numeric values drop the .0 suffix; compound codes kept; empty → "0".
        if 'COD FORNECEDOR' in template.columns:
            def _clean_cod_forn_template(val):
                s = str(val).strip()
                if s in ('nan', '', 'None'):
                    return '0'
                if '/' in s:
                    return s
                try:
                    return str(int(float(s)))
                except (ValueError, TypeError):
                    return s
            template['COD FORNECEDOR'] = template['COD FORNECEDOR'].apply(_clean_cod_forn_template)
        
        
        # Use pattern matching to find latest dated files, with fallback to non-dated versions
        BD_PN = get_latest_file(
            os.path.join(caminho_base, caminho_BD, "BD_CADASTRO_PN_*.xlsx"),
            fallback=os.path.join(caminho_base, caminho_BD, "BD_CADASTRO_PN.xlsx")
        )
        BD_MDR = get_latest_file(
            os.path.join(caminho_base, caminho_BD, "BD_CADASTRO_MDR_*.xlsx"),
            fallback=os.path.join(caminho_base, caminho_BD, "BD_CADASTRO_MDR.xlsx")
        )
        
        if BD_PN is None:
            raise FileNotFoundError("BD_CADASTRO_PN file not found. Please ensure database files are available.")
        if BD_MDR is None:
            raise FileNotFoundError("BD_CADASTRO_MDR file not found. Please ensure database files are available.")
        
        VEÍCULOS = os.path.join(caminho_base,caminho_BD,"VEÍCULOS.xlsx")
        db_empilhamento = os.path.join(caminho_base,caminho_BD,"BD_EMPILHAMENTO_EMBALAGENS.xlsx")
        db_efi = os.path.join(caminho_base,caminho_BD,"BD_CADASTRO_MDR_PERDA_COMPRIMENTO.xlsx")
        PN_CT_path = os.path.join(caminho_base,caminho_BD,"PN_Conta_trabalho.xlsx")
       
        # ------------------Working in the DB structrue------------------
        # Use 'Int64' (nullable integer) instead of 'int' to handle NaN values in the Excel file
        db_PN = pd.read_excel(BD_PN, sheet_name='BD', dtype={'DESENHO': str})
        db_PN = db_PN.rename(columns={'CÓD. FORNECEDOR': 'COD FORNECEDOR'})
        
        # Filter for EMPRESA = 1, 10.12 (not separate 10 and 12!)
        # Note: EMPRESA 10.12 is a single float value in the database
        if 'EMPRESA' in db_PN.columns:
            db_PN = db_PN[db_PN['EMPRESA'].isin([1, 1.0, 10.12,191,191.0])]
        else:
            print("[WARNING] Column 'EMPRESA' not found in BD_CADASTRO_PN")

        db_MDR = pd.read_excel(BD_MDR, sheet_name='BD')
        db_MDR = db_MDR.rename(columns={'DESCRIÇÃO2': 'DESCRIÇÃO'})
        
        # Filter for EMPRESA = 1, 10.12 (not separate 10 and 12!)
        if 'EMPRESA' in db_MDR.columns:
            db_MDR = db_MDR[db_MDR['EMPRESA'].isin([1, 1.0, 10.12,191,191.0])]
        else:
            print("[WARNING] Column 'EMPRESA' not found in BD_CADASTRO_MDR")

        db_veiculos = pd.read_excel(VEÍCULOS, sheet_name='VEÍCULOS')

        db_empilhamento = pd.read_excel(db_empilhamento, sheet_name='BD')
        db_empilhamento = db_empilhamento.rename(columns={'CÓD. FORNECEDOR': 'COD FORNECEDOR'})

        db_efi = pd.read_excel(db_efi,sheet_name='BD')
        
        # --- Normalização de tipos ---
        db_PN['DESENHO ATUALIZAÇÃO'] = pd.to_datetime(db_PN['DESENHO ATUALIZAÇÃO'], errors='coerce')
        db_MDR['VOLUME'] = pd.to_numeric(db_MDR['VOLUME'], errors='coerce')
        db_MDR['MDR PESO'] = pd.to_numeric(db_MDR['MDR PESO'], errors='coerce')
        db_PN['PESO (Kg) MATERIAL'] = pd.to_numeric(db_PN['PESO (Kg) MATERIAL'], errors='coerce')
        
        
        
        db_PN = db_PN.sort_values('DESENHO ATUALIZAÇÃO', ascending=False)
        
        # Criar chave composta DESENHO+MDR em db_PN
        db_PN['KEY'] = db_PN['DESENHO'].astype(str) + '_' + db_PN['MDR'].astype(str)

        # --- Mapeamentos únicos para .map() seguros ---
        # Filter out nan values and keep most recent entries for db_PN mappings
        mapa_fornecedores = db_PN.drop_duplicates('COD FORNECEDOR').set_index('COD FORNECEDOR')['FORNECEDOR']

        # Mapas baseados na chave composta - already sorted by DESENHO ATUALIZAÇÃO descending
        # This ensures we always get the most recent non-null values
        db_PN_valid_desc = db_PN[db_PN['DESCRIÇÃO'].notna()]
        mapa_pn = db_PN_valid_desc.drop_duplicates('KEY', keep='first').set_index('KEY')['DESCRIÇÃO']
        
        mapa_mdr = db_PN.drop_duplicates('KEY', keep='first').set_index('KEY')['MDR']
        
        # For QME and PESO, filter out invalid values before mapping
        db_PN_valid_qme = db_PN[db_PN['QME'].notna() & (db_PN['QME'] > 0)]
        mapa_qme = db_PN_valid_qme.drop_duplicates('KEY', keep='first').set_index('KEY')['QME']
        
        # Peso PN mappings - both composite key (COD FORNECEDOR|KEY) and KEY-only fallback
        db_PN_valid_peso = db_PN[db_PN['PESO (Kg) MATERIAL'].notna()].copy()
        db_PN_valid_peso['CHAVE_PESO_PN'] = db_PN_valid_peso['COD FORNECEDOR'].apply(_codigo_principal).astype(str) + '|' + db_PN_valid_peso['KEY'].astype(str)
        mapa_peso_pn = db_PN_valid_peso.drop_duplicates('CHAVE_PESO_PN', keep='first').set_index('CHAVE_PESO_PN')['PESO (Kg) MATERIAL']
        mapa_peso_pn_fallback = db_PN_valid_peso.drop_duplicates('KEY', keep='first').set_index('KEY')['PESO (Kg) MATERIAL']

        # Mapas vindos do db_MDR - filter out nan values BEFORE creating mappings
        # Get the correct column name for FORNECEDOR
        coluna_fornecedor_mdr = 'CÓD. FORNECEDOR' if 'CÓD. FORNECEDOR' in db_MDR.columns else ('COD FORNECEDOR' if 'COD FORNECEDOR' in db_MDR.columns else None)
        
        db_MDR_valid_desc = db_MDR[db_MDR['DESCRIÇÃO'].notna()]
        mapa_descricao_mdr = db_MDR_valid_desc.drop_duplicates('MDR', keep='first').set_index('MDR')['DESCRIÇÃO']
        
       
       
       
      # Volume mappings - both composite key and MDR-only fallback
        db_MDR_valid_volume = db_MDR[db_MDR['VOLUME'].notna()].copy()

        db_MDR_valid_volume['MDR_CHAVE'] = (
            db_MDR_valid_volume['MDR'].apply(_mdr_chave)
        )

        if coluna_fornecedor_mdr and coluna_fornecedor_mdr in db_MDR_valid_volume.columns:
            db_MDR_valid_volume['CHAVE_VOLUME'] = (
                db_MDR_valid_volume[coluna_fornecedor_mdr]
                .apply(_codigo_principal)
                .astype(str)
                + '|'
                + db_MDR_valid_volume['MDR_CHAVE']
            )

            mapa_volume = (
                db_MDR_valid_volume
                .drop_duplicates('CHAVE_VOLUME', keep='first')
                .set_index('CHAVE_VOLUME')['VOLUME']
            )
        else:
            mapa_volume = {}

        mapa_volume_mdr = (
            db_MDR_valid_volume
            .drop_duplicates('MDR_CHAVE', keep='first')
            .set_index('MDR_CHAVE')['VOLUME']
        )
                        
                
                
                
                
                
       
       
       # Peso MDR mappings - both composite key and MDR-only fallback
        db_MDR_valid_peso = db_MDR[db_MDR['MDR PESO'].notna()].copy()

        if coluna_fornecedor_mdr and coluna_fornecedor_mdr in db_MDR_valid_peso.columns:
            db_MDR_valid_peso['CHAVE_PESO_MDR'] = (
                db_MDR_valid_peso[coluna_fornecedor_mdr]
                .apply(_codigo_principal)
                .astype(str)
                + '|'
                + db_MDR_valid_peso['MDR'].apply(_mdr_chave)
            )

            mapa_peso_mdr = (
                db_MDR_valid_peso
                .drop_duplicates('CHAVE_PESO_MDR', keep='first')
                .set_index('CHAVE_PESO_MDR')['MDR PESO']
            )
        else:
            mapa_peso_mdr = {}

        mapa_peso_mdr_fallback = (
            db_MDR_valid_peso
            .drop_duplicates('MDR', keep='first')
            .set_index('MDR')['MDR PESO']
        )



        
        mapa_peso_max = db_veiculos.set_index('COD VEICULO')['PESO MAXIMO']

       

        # Passo 1: primeiro trazer MDR pelo DESENHO, para podermos montar a KEY
        # Filter out nan MDR values and use most recent (already sorted by DESENHO ATUALIZAÇÃO)
        db_PN_valid_mdr = db_PN[db_PN['MDR'].notna()]
        
        # --- ADDED: PRINTING PN MATCHES TO THE DEBUG CONDITION ---
        if DEBUG_PNS:
            for debug_pn in DEBUG_PNS:
                pn_str = str(debug_pn).strip()
                print(f"\n[DEBUG PN] Verificando PN (DESENHO): {pn_str}")
                in_template = template[template['DESENHO'].astype(str).str.strip() == pn_str]
                if not in_template.empty:
                    print(f"  -> Encontrado {len(in_template)} vez(es) no arquivo de demanda (Template).")
                    in_db = db_PN_valid_mdr[db_PN_valid_mdr['DESENHO'].astype(str).str.strip() == pn_str]
                    if not in_db.empty:
                        print(f"  -> Encontrado no BD_CADASTRO_PN. Matches disponíveis:")
                        for _, row_db in in_db.iterrows():
                            print(f"     * Fornecedor: {row_db.get('COD FORNECEDOR', 'N/A')} | MDR: {row_db['MDR']} | Descrição: {row_db.get('DESCRIÇÃO', 'N/A')}")
                    else:
                        print(f"  -> ERRO: NÃO encontrado no BD_CADASTRO_PN (ou sem MDR válido). O PN não será mapeado para o Viajante!")
                else:
                    print(f"  -> Não encontrado nas demandas atuais do Template.")
        # ---------------------------------------------------------

        template['MDR'] = template['DESENHO'].map(
            db_PN_valid_mdr.drop_duplicates('DESENHO', keep='first').set_index('DESENHO')['MDR']
        )
        
        # Track PNs not found in BD_CADASTRO_PN (no MDR means PN not registered)
        template['PN_NOT_FOUND'] = template['MDR'].isna() | (template['MDR'].astype(str).str.strip() == '')


        # Passo 2: agora que já temos MDR no template, podemos montar a KEY
        template['KEY'] = template['DESENHO'].astype(str) + '_' + template['MDR'].astype(str)
        

        # Passo 3: enriquecer com os mapas
        template['PESO_MAXIMO'] = template['VEÍCULO'].map(mapa_peso_max)
        
        # === NEW LOGIC: Determine which code (COD FORNECEDOR or COD IMS) works for each supplier ===
        # This ensures we use the SAME code consistently for ALL lookups
        def determine_supplier_code(row):
            """
            Try both COD FORNECEDOR and COD IMS to see which one exists in BD_PN.
            Return the code that works, along with which field it came from.
            """
            cod_forn = row.get('COD FORNECEDOR')
            cod_ims = row.get('COD IMS')
            agrupamento = row.get('AGRUPAMENTO')
            
            # Try COD IMS first
            if pd.notna(cod_ims):
                candidatos_ims = _normalizar_codigos_campo(cod_ims)
                for codigo in candidatos_ims:
                    codigo_chave = _codigo_principal(codigo)
                    if codigo_chave:
                        try:
                            codigo_num = float(codigo_chave)
                            if codigo_num in mapa_fornecedores.index:
                                # Found in BD_PN using COD IMS
                                for debug_code in DEBUG_SUPPLIERS:
                                    if str(debug_code) in str(cod_forn) or str(debug_code) in str(cod_ims):
                                        add_debug_info(
                                            debug_code,
                                            "SUPPLIER_CODE_DETERMINATION",
                                            f"Determined supplier code for DESENHO {row.get('DESENHO')}",
                                            {
                                                'AGRUPAMENTO': agrupamento,
                                                'COD FORNECEDOR': cod_forn,
                                                'COD IMS': cod_ims,
                                                'Code that works': codigo_chave,
                                                'Source': 'COD IMS',
                                                'Status': '✓ Found in BD_PN - will use COD IMS for all lookups'
                                            },
                                            agrupamento=agrupamento
                                        )
                                        break
                                return pd.Series({'WORKING_CODE': codigo_chave, 'CODE_SOURCE': 'COD_IMS'})
                        except (ValueError, TypeError):
                            pass
            
            # Try COD FORNECEDOR as fallback
            if pd.notna(cod_forn):
                candidatos_forn = _normalizar_codigos_campo(cod_forn)
                for codigo in candidatos_forn:
                    codigo_chave = _codigo_principal(codigo)
                    if codigo_chave:
                        try:
                            codigo_num = float(codigo_chave)
                            if codigo_num in mapa_fornecedores.index:
                                # Found in BD_PN using COD FORNECEDOR
                                for debug_code in DEBUG_SUPPLIERS:
                                    if str(debug_code) in str(cod_forn) or str(debug_code) in str(cod_ims):
                                        add_debug_info(
                                            debug_code,
                                            "SUPPLIER_CODE_DETERMINATION",
                                            f"Determined supplier code for DESENHO {row.get('DESENHO')}",
                                            {
                                                'AGRUPAMENTO': agrupamento,
                                                'COD FORNECEDOR': cod_forn,
                                                'COD IMS': cod_ims,
                                                'Code that works': codigo_chave,
                                                'Source': 'COD FORNECEDOR',
                                                'Status': '✓ Found in BD_PN - will use COD FORNECEDOR for all lookups'
                                            },
                                            agrupamento=agrupamento
                                        )
                                        break
                                return pd.Series({'WORKING_CODE': codigo_chave, 'CODE_SOURCE': 'COD_FORNECEDOR'})
                        except (ValueError, TypeError):
                            pass
            
            # Neither code found in BD_PN
            for debug_code in DEBUG_SUPPLIERS:
                if str(debug_code) in str(cod_forn) or str(debug_code) in str(cod_ims):
                    add_debug_info(
                        debug_code,
                        "SUPPLIER_CODE_DETERMINATION",
                        f"Could not determine supplier code for DESENHO {row.get('DESENHO')}",
                        {
                            'AGRUPAMENTO': agrupamento,
                            'COD FORNECEDOR': cod_forn,
                            'COD IMS': cod_ims,
                            'Status': '✗ Neither code found in BD_PN - lookups will fail'
                        },
                        agrupamento=agrupamento
                    )
                    break
            
            return pd.Series({'WORKING_CODE': np.nan, 'CODE_SOURCE': None})
        
        # Determine which code works for each row
        template[['WORKING_CODE', 'CODE_SOURCE']] = template.apply(determine_supplier_code, axis=1)

        # ---> ADDED THIS CRITICAL FIX <---
        # Update COD FORNECEDOR to the confirmed WORKING_CODE so all downstream groupings (like df_saturacao) use the right ID
        mask_valid_code = template['WORKING_CODE'].notna()
        template.loc[mask_valid_code, 'COD FORNECEDOR'] = template.loc[mask_valid_code, 'WORKING_CODE'].apply(lambda x: str(int(float(x))) if pd.notna(x) else x)
        
        # Use the working code for FORNECEDOR mapping
        template['WORKING_CODE_NUMERIC'] = pd.to_numeric(template['WORKING_CODE'], errors='coerce')
        template['FORNECEDOR'] = template['WORKING_CODE_NUMERIC'].map(mapa_fornecedores)
       
        # Clean up FORNECEDOR column - remove .0 suffix if it exists (when mapping fails, it might keep numeric values)
        if 'FORNECEDOR' in template.columns:
            template['FORNECEDOR'] = template['FORNECEDOR'].astype(str).str.replace(r'\.0$', '', regex=True)
            # If FORNECEDOR is 'nan', replace with empty string
            template['FORNECEDOR'] = template['FORNECEDOR'].replace('nan', '')
       
        template['DESCRIÇÃO MATERIAL'] = template['KEY'].map(mapa_pn)
        template['MDR'] = template['KEY'].map(mapa_mdr)  # reforça MDR correto do KEY
        
        template['DESCRIÇÃO DA EMBALAGEM'] = template['MDR'].map(mapa_descricao_mdr)
        
        template['QME'] = template['KEY'].map(mapa_qme)

        # Ensure QME is valid (not zero, not NaN) before division
        # template['QME'] = template['QME'].fillna(1)  # Replace NaN with 1 to avoid division issues
        # template['QME'] = template['QME'].replace(0, 1)  # Replace 0 with 1 to avoid division by zero
        mask_invalid = (
                template['QME'].isna() |
                (template['QME'] == 0) |
                (~np.isfinite(template['QME']))
            )

        template['QTD EMBALAGENS'] = np.where(
                mask_invalid,
                0,
                np.ceil(template['QTDE'] / template['QME'])
            )
                    
        
        # Clean up any infinity values in QTD EMBALAGENS
        template['QTD EMBALAGENS'] = template['QTD EMBALAGENS'].replace([np.inf, -np.inf], np.nan).fillna(0)

        # === RESOLVER FUNCTIONS - Create intermediate Series for calculations ===
        # Use COD IMS + MDR as the primary volume key, trying every IMS code before
        # falling back to COD FORNECEDOR and, if needed, MDR-only.
        def resolver_volume(row):
            mdr_chave = _mdr_chave(row.get('MDR'))
            candidatos = _normalizar_codigos_campo(row.get('COD IMS'))

            if not candidatos:
                candidatos = _normalizar_codigos_campo(row.get('COD FORNECEDOR'))

            for codigo in candidatos:
                chave = _chave_fornecedor_mdr(codigo, mdr_chave)
                if chave:
                    volume = mapa_volume.get(chave, np.nan)
                    if pd.notna(volume):
                        return pd.Series({'VOLUME_UNITARIO': volume, 'VOLUME_KEY': chave})

            volume_fallback = mapa_volume_mdr.get(mdr_chave, np.nan)
            return pd.Series({'VOLUME_UNITARIO': volume_fallback, 'VOLUME_KEY': ''})

        volume_info = template.apply(resolver_volume, axis=1)
        
        # Handle empty volume_info (when template is empty)
        if volume_info.empty or 'VOLUME_UNITARIO' not in volume_info.columns:
            adicionar_erro("Não foi possível calcular volume - dados insuficientes", "ERRO")
            template['M³'] = 0
            volume_por_mdr = pd.Series([0] * len(template), index=template.index)
            volume_lookup_chave = pd.Series([''] * len(template), index=template.index)
        else:
            volume_por_mdr = volume_info['VOLUME_UNITARIO'].replace([np.inf, -np.inf], np.nan).fillna(0)
            volume_lookup_chave = volume_info['VOLUME_KEY']
        
        # Only calculate M³ if volume_info was valid
        if not volume_info.empty and 'VOLUME_UNITARIO' in volume_info.columns:
            template['M³'] = round(template['QTD EMBALAGENS'] * volume_por_mdr, 3)
        
        # Use COD IMS + KEY as the primary peso material key, trying every IMS code before
        # falling back to COD FORNECEDOR and, if needed, KEY-only.
        def resolver_peso_material(row):
            key = row.get('KEY')
            if pd.isna(key) or str(key).strip() == '':
                return np.nan
            
            candidatos = _normalizar_codigos_campo(row.get('COD IMS'))

            if not candidatos:
                candidatos = _normalizar_codigos_campo(row.get('COD FORNECEDOR'))

            for codigo in candidatos:
                codigo_chave = _codigo_principal(codigo)
                if codigo_chave:
                    chave_completa = f"{codigo_chave}|{key}"
                    peso = mapa_peso_pn.get(chave_completa, np.nan)
                    if pd.notna(peso):
                        return peso

            peso_fallback = mapa_peso_pn_fallback.get(key, np.nan)
            return peso_fallback

        peso_material_unitario = template.apply(resolver_peso_material, axis=1)
        template['PESO MAT'] = round(template['QTDE'] * peso_material_unitario, 1)
        
        # Use COD IMS + MDR as the primary peso key, trying every IMS code before
        # falling back to COD FORNECEDOR and, if needed, MDR-only.
        def resolver_peso_mdr(row):
            mdr_chave = _mdr_chave(row.get('MDR'))
            candidatos = _normalizar_codigos_campo(row.get('COD IMS'))

            if not candidatos:
                candidatos = _normalizar_codigos_campo(row.get('COD FORNECEDOR'))

            for codigo in candidatos:
                chave = _chave_fornecedor_mdr(codigo, mdr_chave)
                if chave:
                    peso = mapa_peso_mdr.get(chave, np.nan)
                    if pd.notna(peso):
                        return peso

            peso_fallback = mapa_peso_mdr_fallback.get(mdr_chave, np.nan)
            return peso_fallback

        peso_mdr_por_embalagem = template.apply(resolver_peso_mdr, axis=1)
        template['PESO MDR'] = round(template['QTD EMBALAGENS'] * peso_mdr_por_embalagem, 1)
        
        # Garante que NaN e infinitos sejam tratados como 0 antes de somar
        template['M³'] = template['M³'].replace([np.inf, -np.inf], np.nan).fillna(0)
        template['PESO MAT'] = template['PESO MAT'].replace([np.inf, -np.inf], np.nan).fillna(0)
        template['PESO MDR'] = template['PESO MDR'].replace([np.inf, -np.inf], np.nan).fillna(0)
        template['PESO TOTAL'] = round(template['PESO MAT'] + template['PESO MDR'], 1)


        # ---> REWRITTEN THIS DEBUG MASK FIX <---
        # The previous code had a hardcoded `debug_fornecedor = ''` which broke this print completely
        debug_mask = pd.Series(False, index=template.index)
        for debug_code in DEBUG_SUPPLIERS:
            if 'COD FORNECEDOR' in template.columns:
                debug_mask = debug_mask | template['COD FORNECEDOR'].astype(str).str.contains(str(debug_code))
            if 'COD IMS' in template.columns:
                debug_mask = debug_mask | template['COD IMS'].astype(str).str.contains(str(debug_code))
        
        if DEBUG_PNS:
            for dpn in DEBUG_PNS:
                debug_mask = debug_mask | (template['DESENHO'].astype(str).str.strip() == str(dpn).strip())
        
        if debug_mask.any():
            print(f"\n[DEBUG] Analisando linhas finais antes de salvar no Template.xlsx")
            for idx, dbg_row in template.loc[debug_mask].iterrows():
                # Volume info
                volume_mdr = volume_por_mdr.loc[idx]
                chave_volume = volume_lookup_chave.loc[idx]
                volume_txt = f"{float(volume_mdr):.3f}" if pd.notna(volume_mdr) else "N/A"
                qtd_emb = dbg_row['QTD EMBALAGENS']
                m3_txt = f"{float(dbg_row['M³']):.3f}" if pd.notna(dbg_row['M³']) else "N/A"
                
                # Build keys for display
                cod_ims = dbg_row.get('COD IMS')
                cod_forn = dbg_row.get('COD FORNECEDOR')
                key = dbg_row.get('KEY')
                candidatos = _normalizar_codigos_campo(cod_ims)
                if not candidatos:
                    candidatos = _normalizar_codigos_campo(cod_forn)
                
                # QME key
                qme_key_used = f"{_codigo_principal(candidatos[0]) if candidatos else ''}|{key}" if key else "N/A"
                
                # PESO MATERIAL key (COD FORNECEDOR + KEY)
                peso_mat_key_used = f"{_codigo_principal(candidatos[0]) if candidatos else ''}|{key}" if key else "N/A"
                
                # Get PESO MATERIAL unit from PN file (using composite key)
                peso_mat_unit_pn = peso_material_unitario.loc[idx]
                peso_mat_unit_txt = f"{float(peso_mat_unit_pn):.4f}" if pd.notna(peso_mat_unit_pn) else "N/A"
                qtde = dbg_row['QTDE']
                peso_mat_calc = f"{qtde} x {peso_mat_unit_txt}" if pd.notna(peso_mat_unit_pn) else "N/A"
                peso_mat_result = f"{float(dbg_row['PESO MAT']):.1f}" if pd.notna(dbg_row['PESO MAT']) else "N/A"
                
                # Build PESO MDR key for display (COD FORNECEDOR + MDR)
                mdr = dbg_row.get('MDR')
                peso_mdr_key_used = _chave_fornecedor_mdr(candidatos[0] if candidatos else '', mdr)
                peso_mdr_unit = peso_mdr_por_embalagem.loc[idx]
                peso_mdr_unit_txt = f"{float(peso_mdr_unit):.2f}" if pd.notna(peso_mdr_unit) else "N/A"
                peso_mdr_calc = f"{qtd_emb} x {peso_mdr_unit_txt}" if pd.notna(peso_mdr_unit) else "N/A"
                peso_mdr_result = f"{float(dbg_row['PESO MDR']):.1f}" if pd.notna(dbg_row['PESO MDR']) else "N/A"
                
                # PESO TOTAL calculation
                peso_total_calc = f"{peso_mat_result} + {peso_mdr_result}"
                peso_total_result = f"{float(dbg_row['PESO TOTAL']):.1f}" if pd.notna(dbg_row['PESO TOTAL']) else "N/A"
                
                print(
                    f"  VOLUME_KEY={chave_volume or _chave_fornecedor_mdr(dbg_row.get('COD IMS'), dbg_row.get('MDR'))} | "
                    f"QME_KEY={qme_key_used} | PESO_MAT_KEY={peso_mat_key_used} | PESO_MDR_KEY={peso_mdr_key_used or 'N/A'}"
                )
                print(
                    f"  DESENHO={dbg_row['DESENHO']} | MDR={dbg_row['MDR']} | QME={dbg_row['QME']} | QTDE={qtde}"
                )
                print(
                    f"  VOLUME={volume_txt} | QTD_EMB={qtd_emb} | M³ = {qtd_emb} x {volume_txt} = {m3_txt}"
                )
                print(
                    f"  PESO_MAT_UNIT_PN={peso_mat_unit_txt} kg | PESO_MAT = {peso_mat_calc} = {peso_mat_result} kg"
                )
                print(
                    f"  PESO_MDR_UNIT={peso_mdr_unit_txt} kg | PESO_MDR = {peso_mdr_calc} = {peso_mdr_result} kg"
                )
                print(
                    f"  PESO_TOTAL = {peso_total_calc} = {peso_total_result} kg"
                )
                print()

        # Final cleanup: re-apply string normalisation in case any operation reintroduced floats
        if 'COD FORNECEDOR' in template.columns:
            def _clean_cod_forn_template(val):
                s = str(val).strip()
                if s in ('nan', '', 'None'):
                    return '0'
                if '/' in s:
                    return s
                try:
                    return str(int(float(s)))
                except (ValueError, TypeError):
                    return s
            template['COD FORNECEDOR'] = template['COD FORNECEDOR'].apply(_clean_cod_forn_template)
        
       
        # Ensure COD FLUXO exists (older Template.xlsx files may not have it)
        if 'COD FLUXO' not in template.columns:
            template['COD FLUXO'] = None

        if 'DATA COLETA' not in template.columns:
            template['DATA COLETA'] = None

        template = template[['AGRUPAMENTO', 'DATA COLETA', 'COD IMS','COD FORNECEDOR', 'FORNECEDOR', 'DESENHO', 'QTDE','PLANTA', 'DESCRIÇÃO MATERIAL',
                             'MDR', 'DESCRIÇÃO DA EMBALAGEM', 'QME', 'QTD EMBALAGENS', 'TIPO SATURACAO',
                             'VEÍCULO', 'M³', 'PESO MAT', 'PESO MDR', 'PESO TOTAL', 'PESO_MAXIMO']]
        
        # --- Remove duplicates BEFORE calculations ---
        # Deduplicate after enrichment but before saturação calculations
        # This ensures all calculations (volume, peso, ocupação) are done on clean data
       
        template = template.drop_duplicates(subset=['AGRUPAMENTO', 'DATA COLETA', 'COD FORNECEDOR', 'PLANTA', 'DESENHO', 'QTDE', 'VEÍCULO', 'TIPO SATURACAO']).reset_index(drop=True)
    
       
        # --- Construção da aba Saturação ---
        # Group by COD FLUXO (unique per fluxo row in FLUXO.xlsx) so each fluxo is
        # calculated independently — quantities from different fluxos are never summed.
        df_saturacao = (
            template.groupby(['AGRUPAMENTO', 'COD FORNECEDOR', 'MDR', 'VEÍCULO'], as_index=False)['QTD EMBALAGENS']
            .sum()
            .rename(columns={'MDR': 'EMBALAGEM', 'QTD EMBALAGENS': 'TOTAL DE CXS'})
        )

        # Create mappings from db_MDR - filter out NaN values before deduplication
        # to ensure we don't get empty/null values when valid values exist
        db_MDR_valid_paletizavel = db_MDR[db_MDR['CAIXA PLÁSTICA'].notna()]
        mapa_paletizavel = db_MDR_valid_paletizavel.drop_duplicates('MDR').set_index('MDR')['CAIXA PLÁSTICA']
        
        # For CAIXAS POR PALLET, filter out NaN and use the most common value (mode)
        # This prevents picking the first row if it has NaN when other rows have valid values
        db_MDR_valid_cxs = db_MDR[db_MDR['CAIXAS POR PALLET'].notna()]
        
        # Group by MDR and take the mode (most common value) for CAIXAS POR PALLET
        # If multiple modes exist, take the first one
        mapa_cxs_por_pallet = db_MDR_valid_cxs.groupby('MDR')['CAIXAS POR PALLET'].agg(
            lambda x: x.mode()[0] if not x.mode().empty else x.iloc[0]
        )

        df_saturacao['CX_PALETIZÁVEL'] = df_saturacao['EMBALAGEM'].map(mapa_paletizavel).fillna(0).astype(int)
        df_saturacao['CXS_POR_PALLET'] = df_saturacao.apply(
            lambda row: 1 if row['CX_PALETIZÁVEL'] != 1 else (
                mapa_cxs_por_pallet.get(row['EMBALAGEM'], 1) or 1), axis=1
        )
        
        df_saturacao['CXS/PALLETS_TOTAL'] = df_saturacao['TOTAL DE CXS'] / df_saturacao['CXS_POR_PALLET']

        # Mapeia de código do veículo (ex: 4) -> coluna de capacidade no db_MDR (ex: "14 x 2,4 x 2,78")
        mapa_coluna_capacidade = db_veiculos.set_index('COD VEICULO')['VEICULOS'].to_dict()
        

        # Garante que os MDRs na base estejam em caixa alta
        db_MDR['MDR'] = db_MDR['MDR'].astype(str).str.upper()

        def obter_capacidade_por_linha(row):
            try:
                mdr = str(row['EMBALAGEM']).upper()
                cod_veic = row['VEÍCULO']
                fornecedor = row['COD FORNECEDOR']
                coluna = mapa_coluna_capacidade.get(cod_veic)

                if not coluna or not isinstance(coluna, str):
                    return np.nan
                if coluna not in db_MDR.columns:
                    return np.nan

                # Try supplier-specific lookup first
                supplier_codes = [fornecedor]
                
                # Use exact matching
                filtro_fornecedor = (db_MDR['MDR'] == mdr) & (db_MDR['CÓD. FORNECEDOR'].isin(supplier_codes))
                capacidade_data_forn = db_MDR.loc[filtro_fornecedor, coluna]
                
                # Ensure we got a Series
                if isinstance(capacidade_data_forn, pd.DataFrame):
                    return np.nan
                    
                capacidade_series_forn = capacidade_data_forn.dropna()
                
                if not capacidade_series_forn.empty:
                    # Found supplier-specific capacity - use HYBRID approach
                    capacity_mean = float(capacidade_series_forn.mean())
                    capacity_min = float(capacidade_series_forn.min())
                    capacity_max = float(capacidade_series_forn.max())
                    mode_result = capacidade_series_forn.mode()
                    capacity_mode = float(mode_result.iloc[0]) if not mode_result.empty else capacity_mean
                    
                    # Hybrid: Use MODE if high variance, otherwise MAX
                    variance = capacity_max - capacity_min
                    if variance > capacity_mode * 0.5:
                        return capacity_mode
                    else:
                        return capacity_max
                
                # Fall back to all suppliers for this MDR, use MAX
                filtro = db_MDR['MDR'] == mdr
                capacidade_data = db_MDR.loc[filtro, coluna]
                
                # Ensure we got a Series
                if isinstance(capacidade_data, pd.DataFrame):
                    return np.nan
                    
                capacidade_series = capacidade_data.dropna()

                if capacidade_series.empty:
                    return np.nan

                # Use MAX approach
                capacity_mean = float(capacidade_series.mean())
                capacity_min = float(capacidade_series.min())
                capacity_max = float(capacidade_series.max())
                mode_result = capacidade_series.mode()
                capacity_mode = float(mode_result.iloc[0]) if not mode_result.empty else capacity_mean
                
                return capacity_max
                
            except Exception as e:
                print(f"[WARNING] Error in obter_capacidade_por_linha: {e}")
                return np.nan

        df_saturacao['CAPACIDADE'] = df_saturacao.apply(obter_capacidade_por_linha, axis=1)
        df_saturacao['VEÍCULO'] = df_saturacao['VEÍCULO'].fillna(0)
        df_saturacao['VEÍCULO'] = df_saturacao['VEÍCULO'].astype(int)
        
        # Converte para numérico, tratando valores não numéricos
        df_saturacao['CAPACIDADE'] = pd.to_numeric(df_saturacao['CAPACIDADE'], errors='coerce')
        df_saturacao['CXS/PALLETS_TOTAL'] = pd.to_numeric(df_saturacao['CXS/PALLETS_TOTAL'], errors='coerce')

        # Map COD FORNECEDOR to FORNECEDOR for matching with empilhamento database
        df_saturacao['FORNECEDOR'] = df_saturacao['COD FORNECEDOR'].map(mapa_fornecedores)
        
        bases = set(zip(db_empilhamento['FORNECEDOR'], db_empilhamento['MDR BASE']))
        sobrepostas = set(zip(db_empilhamento['FORNECEDOR'], db_empilhamento['MDR SOBREPOSTA']))
        df_saturacao['EMBALAGEM_BASE'] = df_saturacao.apply(
            lambda row: 1 if (row['FORNECEDOR'], row['EMBALAGEM']) in bases else 0, axis=1)
        df_saturacao['EMBALAGEM_SOBREPOSTA'] = df_saturacao.apply(
            lambda row: 1 if (row['FORNECEDOR'], row['EMBALAGEM']) in sobrepostas else 0, axis=1)

        df_saturacao['CHAVE'] = df_saturacao['COD FORNECEDOR'].astype(str) + '-' + df_saturacao['EMBALAGEM'].astype(str)


        # --- Eficiência de empilhamento por embalagem ---
        # Always use per-row vehicle lookup (vehicle is determined from Excel file)
        def obter_eficiencia_por_linha(row):
            chave = str(row['COD FORNECEDOR']) + '-' + str(row['EMBALAGEM'])
            cod_veic = row['VEÍCULO']  # Fixed: was 'VEICULO', should be 'VEÍCULO' with accent
            coluna_veic = mapa_coluna_capacidade.get(cod_veic)
            
            if not coluna_veic or coluna_veic not in db_efi.columns:
                return 1.0  # Default efficiency
            
            # Filter db_efi for this CHAVE and get efficiency from appropriate vehicle column
            filtro = db_efi['CHAVE FORNE + MDR'] == chave
            efi_series = db_efi.loc[filtro, coluna_veic].dropna()
            
            if efi_series.empty:
                return 1.0
            
            return efi_series.values[0]
        
        df_saturacao['EFICIÊNCIA_COMPRIMENTO'] = df_saturacao.apply(obter_eficiencia_por_linha, axis=1)
        df_saturacao['EFICIÊNCIA_COMPRIMENTO'] = df_saturacao['EFICIÊNCIA_COMPRIMENTO'].fillna(1)


        mapa_volume_efi = db_MDR.drop_duplicates('CHAVE EMBALAGENS').set_index('CHAVE EMBALAGENS')['VOLUME']
        df_saturacao['M³ POR EMBALAGEM'] = df_saturacao['CHAVE'].map(mapa_volume_efi) * \
                                            df_saturacao['CXS_POR_PALLET'] * df_saturacao['CXS/PALLETS_TOTAL']

        # --- Cálculo de empilhamento ---
        df_calculo_empilhamento = calcular_empilhamento(df_saturacao, db_empilhamento)

        # --- Saturação final por embalagem ---
        def integrar_saturacao_total(df_sat, df_emp):
            def calcular(row):
                filtro = (df_emp['FORNECEDOR'] == row['COD FORNECEDOR']) & \
                         (df_emp['EMBALAGEM_BASE'] == row['EMBALAGEM'])
                soma_saturacoes = df_emp[filtro]['SATURAÇÃO'].sum()
                # Prevent division by zero
                if pd.isna(row['CAPACIDADE']) or row['CAPACIDADE'] == 0:
                    return 0
                proporcao = row['CXS/PALLETS_TOTAL'] / row['CAPACIDADE']
                result = (proporcao + soma_saturacoes) * row['EFICIÊNCIA_COMPRIMENTO']
                

                
                return result

            df_sat['SATURAÇÃO_TOTAL'] = df_sat.apply(calcular, axis=1)
            # Clean up infinity values
            df_sat['SATURAÇÃO_TOTAL'] = df_sat['SATURAÇÃO_TOTAL'].replace([np.inf, -np.inf], np.nan).fillna(0)
            
            # Prevent division by zero for SATURAÇÃO_POR_MDR
            df_sat['SATURAÇÃO_POR_MDR'] = 0.0
            mask = (df_sat['TOTAL DE CXS'].notna()) & (df_sat['TOTAL DE CXS'] > 0)
            df_sat.loc[mask, 'SATURAÇÃO_POR_MDR'] = df_sat.loc[mask, 'SATURAÇÃO_TOTAL'] / df_sat.loc[mask, 'TOTAL DE CXS']
            df_sat['SATURAÇÃO_POR_MDR'] = df_sat['SATURAÇÃO_POR_MDR'].replace([np.inf, -np.inf], np.nan).fillna(0)
            
            return df_sat

        if not df_calculo_empilhamento.empty:
            df_saturacao = integrar_saturacao_total(df_saturacao, df_calculo_empilhamento)
        else:
            # Prevent division by zero
            df_saturacao['SATURAÇÃO_TOTAL'] = 0.0
            mask = (df_saturacao['CAPACIDADE'].notna()) & (df_saturacao['CAPACIDADE'] > 0)
            df_saturacao.loc[mask, 'SATURAÇÃO_TOTAL'] = df_saturacao.loc[mask, 'CXS/PALLETS_TOTAL'] / df_saturacao.loc[mask, 'CAPACIDADE']
            df_saturacao['SATURAÇÃO_TOTAL'] = df_saturacao['SATURAÇÃO_TOTAL'].replace([np.inf, -np.inf], np.nan).fillna(0)
            
            df_saturacao['SATURAÇÃO_POR_MDR'] = 0.0
            mask = (df_saturacao['TOTAL DE CXS'].notna()) & (df_saturacao['TOTAL DE CXS'] > 0)
            df_saturacao.loc[mask, 'SATURAÇÃO_POR_MDR'] = df_saturacao.loc[mask, 'SATURAÇÃO_TOTAL'] / df_saturacao.loc[mask, 'TOTAL DE CXS']
            df_saturacao['SATURAÇÃO_POR_MDR'] = df_saturacao['SATURAÇÃO_POR_MDR'].replace([np.inf, -np.inf], np.nan).fillna(0)

        # --- Cálculo da SAT por linha ---
        template.loc[:, 'CHAVE'] = template['COD FORNECEDOR'].astype(str) + '-' + template['MDR'].astype(str)
        
        # Merge on both CHAVE and COD FLUXO so each fluxo gets its own saturation
        template = template.merge(df_saturacao[['CHAVE', 'AGRUPAMENTO','SATURAÇÃO_POR_MDR']], on=['CHAVE', 'AGRUPAMENTO'], how='left')
        
        # Clean up SATURAÇÃO_POR_MDR to avoid infinity values
        template['SATURAÇÃO_POR_MDR'] = template['SATURAÇÃO_POR_MDR'].replace([np.inf, -np.inf], np.nan).fillna(0)
        
        template['SAT VOLUME (%)'] = round(template['QTD EMBALAGENS'] * template['SATURAÇÃO_POR_MDR'] * 100, 2)
        
        # Prevent division by zero for SAT PESO calculation
        template['PESO_MAXIMO'] = template['PESO_MAXIMO'].replace(0, np.nan)
        template['SAT PESO (%)'] = round(template['PESO TOTAL'] / template['PESO_MAXIMO'] * 100, 2)
        
        # Clean up any infinity/NaN values in SAT columns
        template['SAT VOLUME (%)'] = template['SAT VOLUME (%)'].replace([np.inf, -np.inf], np.nan).fillna(0)
        template['SAT PESO (%)'] = template['SAT PESO (%)'].replace([np.inf, -np.inf], np.nan).fillna(0)
        
        
       
        template.drop(columns=['CHAVE', 'SATURAÇÃO_POR_MDR'], inplace=True)
        df_saturacao.drop(columns=['CHAVE'], inplace=True)

        # --- Criação das variáveis para a tabela final ---
        # Filter to only include rows that should be in calculations
        template_calc = template[template['INCLUDE_IN_CALC']] if 'INCLUDE_IN_CALC' in template.columns else template
        
        ocupacao = template_calc['SAT VOLUME (%)'].sum()
        qtd_veiculos = (ceil(ocupacao / 100))
        volume = template_calc['M³'].sum()
        peso = template_calc['PESO TOTAL'].sum()
        
        # Handle infinity and NaN values in QTD EMBALAGENS before summing
        embalagens_series = template_calc['QTD EMBALAGENS'].replace([np.inf, -np.inf], np.nan).fillna(0)
        embalagens = embalagens_series.sum()

        # Preenche a tree_resumo (que deve ser passada como argumento)
        resumo_dados = [
            ("Ocupação Total", f"{ocupacao:.2f}%"),
            ("Qtd Veículos", qtd_veiculos),
            ("Volume Total", f"{volume:.1f} m³"),
            ("Peso Total", f"{peso:.1f} kg"),
            #("Peso Máximo", f"{peso_maximo:.1f} kg"),
            ("Embalagens", int(embalagens) if np.isfinite(embalagens) else 0),
        ]

        linhas_validas = template[
            (template['DESENHO'].notna()) &
            (template['COD FORNECEDOR'].notna()) &
            (template['QTDE'] > 0)
            ].shape[0]

        linha_qme = template[
            (template['QME'] > 0) &
            (template['QTDE'] > 0)
            ].shape[0]

        # Limpa e atualiza a tabela tree_resumo
        tree_resumo.delete(*tree_resumo.get_children())
        for item in resumo_dados:
            tree_resumo.insert("", END, values=item)


        # --- Atualiza TreeView (Tkinter) with deduplicated data ---
        tree.delete(*tree.get_children())
        tree["columns"] = list(template.columns)
        tree["show"] = "headings"
        
        # Define custom widths for specific columns - optimized for scrolling
        column_widths = {
            'COD FORNECEDOR': 110,
            'FORNECEDOR': 150,
            'DESENHO': 90,
            'QTDE': 70,
            'DESCRIÇÃO MATERIAL': 200,
            'MDR': 70,
            'DESCRIÇÃO DA EMBALAGEM': 160,
            'QME': 60,
            'QTD EMBALAGENS': 110,
            'TIPO SATURACAO': 90,
            'VEICULO': 70,
            'MOT': 60,
            'M³': 60,
            'PESO MAT': 90,
            'PESO MDR': 90,
            'PESO TOTAL': 100,
            'PESO_MAXIMO': 110,
            'SAT VOLUME (%)': 110,
            'SAT PESO (%)': 100,
            'COD IMS': 90,
            'INCLUDE_IN_CALC': 110
        }
        
        for col in template.columns:
            tree.heading(col, text=col)
            width = column_widths.get(col, 90)  # Default 90 if not specified
            tree.column(col, width=width, anchor="center", stretch=False)
        for _, row in template.iterrows():
            tree.insert("", END, values=list(row))

        desenhar_caminhoes(canvas_caminhoes, ocupacao, caminhao_img)
        
        # --- Second deduplication check before Excel write ---
        # Safety check in case any operations after calculations reintroduced duplicates
        duplicates_before_write = len(template)
        template = template.drop_duplicates(subset=['AGRUPAMENTO', 'DATA COLETA', 'COD FORNECEDOR', 'PLANTA', 'DESENHO', 'QTDE', 'VEÍCULO', 'TIPO SATURACAO']).reset_index(drop=True)
        duplicates_removed_write = duplicates_before_write - len(template)
        if duplicates_removed_write > 0:
            adicionar_erro(f"{duplicates_removed_write} linha(s) duplicada(s) removida(s) antes de salvar VIAJANTE", "INFO")
            print(f"[INFO] Removed {duplicates_removed_write} additional duplicates before Excel write")
        
        
        # --- Exporta para Excel formatado ---
        with pd.ExcelWriter('VIAJANTE.xlsx', engine='openpyxl') as writer:
            template.to_excel(writer, sheet_name='Template Completo', index=False)
            df_saturacao.to_excel(writer, sheet_name='Saturação', index=False)
            df_calculo_empilhamento.to_excel(writer, sheet_name='Calculo Empilhamento', index=False)

            header_fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid')
            header_font = Font(bold=True, color='000000')
            header_align = Alignment(horizontal='center', vertical='center')

            for sheet_name in ['Template Completo', 'Saturação', 'Calculo Empilhamento']:
                ws = writer.sheets[sheet_name]
                for col_num, col in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
                    largura = max(len(str(cell.value) or '') for cell in col) + 2
                    ws.column_dimensions[get_column_letter(col_num)].width = largura
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_align

            if 'MDR' in template.columns:
                # Identify PNs with issues
                pn_nao_cadastrados_list = []
                
                # 1. PNs not in BD_CADASTRO_PN (no MDR)
                pn_sem_mdr = template[
                    template['MDR'].isna() | (template['MDR'].astype(str).str.strip() == '')
                ].copy()
                if not pn_sem_mdr.empty:
                    pn_sem_mdr['MOTIVO'] = 'PN não encontrado no BD_CADASTRO_PN (sem MDR)'
                    pn_nao_cadastrados_list.append(pn_sem_mdr)
                
                # 2. PNs without QME (critical for calculations)
                pn_sem_qme = template[
                    template['MDR'].notna() & 
                    (template['QME'].isna() | (template['QME'] == 0))
                ].copy()
                if not pn_sem_qme.empty:
                    pn_sem_qme['MOTIVO'] = 'PN sem QME cadastrado'
                    pn_nao_cadastrados_list.append(pn_sem_qme)
                
                # 3. PNs without DESCRIÇÃO MATERIAL
                pn_sem_desc = template[
                    template['MDR'].notna() & 
                    (template['DESCRIÇÃO MATERIAL'].isna() | (template['DESCRIÇÃO MATERIAL'].astype(str).str.strip() == ''))
                ].copy()
                if not pn_sem_desc.empty:
                    pn_sem_desc['MOTIVO'] = 'PN sem DESCRIÇÃO MATERIAL cadastrada'
                    pn_nao_cadastrados_list.append(pn_sem_desc)
                
                # Combine all issues
                if pn_nao_cadastrados_list:
                    pn_nao_cadastrados = pd.concat(pn_nao_cadastrados_list, ignore_index=True)
                    
                    # Select columns to keep
                    cols_to_keep = ['AGRUPAMENTO','DATA COLETA','COD IMS','COD FORNECEDOR', 'FORNECEDOR', 'DESENHO', 'QTDE', 'MDR', 'QME','VEÍCULO','MOTIVO']
                    existing_cols = [c for c in cols_to_keep if c in pn_nao_cadastrados.columns]
                    
                    if existing_cols:
                        pn_nao_cadastrados = pn_nao_cadastrados[existing_cols]
                        pn_nao_cadastrados.drop_duplicates(subset=["DESENHO", "MOTIVO"], inplace=True)
                        pn_nao_cadastrados.to_excel(writer, sheet_name='PN Não Cadastrados', index=False)
                        
                        # Log de PNs não cadastrados
                        qtd_pn_faltando = len(pn_nao_cadastrados)
                        qtd_sem_mdr = len(pn_nao_cadastrados[pn_nao_cadastrados['MOTIVO'].str.contains('sem MDR', na=False)])
                        adicionar_erro(f"{qtd_pn_faltando} PN(s) com problemas de cadastro ({qtd_sem_mdr} sem MDR). Verifique a aba 'PN Não Cadastrados' do viajante", "AVISO")
                        
                        # Append to tracking file
                        tracking_file = os.path.join(caminho_base, caminho_BD, 'PNs_Nao_Cadastrados_Log.xlsx')
                        try:
                            if os.path.exists(tracking_file):
                                existing_log = pd.read_excel(tracking_file)
                                # Add timestamp if not present
                                if 'DATA_SOLICITACAO' not in pn_nao_cadastrados.columns:
                                    pn_nao_cadastrados['DATA_SOLICITACAO'] = pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')
                                combined_log = pd.concat([existing_log, pn_nao_cadastrados], ignore_index=True)
                                combined_log.drop_duplicates(subset=['DESENHO', 'COD FORNECEDOR'], keep='last', inplace=True)
                                combined_log.to_excel(tracking_file, index=False)
                            else:
                                pn_nao_cadastrados['DATA_SOLICITACAO'] = pd.Timestamp.now().strftime('%Y-%m-%d %H:%M:%S')
                                pn_nao_cadastrados.to_excel(tracking_file, index=False)
                            print(f"[INFO] PNs não cadastrados salvos em: {tracking_file}")
                        except Exception as e:
                            adicionar_erro(f"Erro ao salvar log de PNs não cadastrados: {str(e)}", "AVISO")
        
        # === Print debug summary at the end of processing ===
        print_debug_summary()

    except Exception as e:
        adicionar_erro(f"Erro crítico ao processar informações: {str(e)}", "ERRO")
        print(f"Erro: {e}")
        traceback.print_exc()


def consolidar_dados(use_manual=False, manual_veiculo=None):
    template = pd.read_excel('VIAJANTE.xlsx', sheet_name='Template Completo')

    # Filtra linhas com quantidade válida e prepara as colunas
    template = template[template['QTDE'] > 0].copy()
    
    # Ensure columns exist (for backward compatibility)
    if 'COD IMS' not in template.columns:
        template['COD IMS'] = ""
    if 'DATA COLETA' not in template.columns:
        template['DATA COLETA'] = ""
    if 'PLANTA' not in template.columns:
        template['PLANTA'] = ""
    
    # Clean up COD FORNECEDOR - remove .0 suffix before converting to string
    template['COD FORNECEDOR'] = template['COD FORNECEDOR'].astype(str).str.replace(r'\.0$', '', regex=True).str.strip()
    
    if 'AGRUPAMENTO' not in template.columns:
        template['AGRUPAMENTO'] = "Sem Agrupamento"
        
    template['AGRUPAMENTO'] = template['AGRUPAMENTO'].astype(str).str.strip()
    template['AGRUPAMENTO'] = template['AGRUPAMENTO'].str.replace(r'\.0$', '', regex=True)
    
    # Clean up FORNECEDOR - remove .0 suffix and handle NaN
    template['FORNECEDOR'] = template['FORNECEDOR'].fillna('').astype(str).str.replace(r'\.0$', '', regex=True)

    # If user forced a manual vehicle, override the template VEICULO column
    
    # Resolve 'VEÍCULO' or 'VEICULO' column name
    veiculo_col = 'VEICULO' if 'VEICULO' in template.columns else ('VEÍCULO' if 'VEÍCULO' in template.columns else None)

    # Build vehicle code -> display name mapping (if BD/VEÍCULOS exists)
    code_to_vehicle_name = {}
    try:
        veic_path = os.path.join(caminho_base, 'BD', 'VEÍCULOS.xlsx')
        if os.path.exists(veic_path):
            db_veic = pd.read_excel(veic_path)
            cols = {c.strip().upper(): c for c in db_veic.columns}
            code_col = None
            name_col = None
            for up, orig in cols.items():
                if up == 'DESCRICAO':
                    name_col = orig
                if 'COD' in up and 'VEIC' in up:
                    code_col = orig
            if code_col is None:
                for up, orig in cols.items():
                    if 'COD' in up:
                        code_col = orig
                        break
            if name_col is None:
                for up, orig in cols.items():
                    if 'VEIC' in up or 'VEICULO' in up:
                        name_col = orig
                        break
            if name_col is None and len(db_veic.columns) > 1:
                name_col = db_veic.columns[1]

            if code_col and name_col:
                for _, r in db_veic.iterrows():
                    raw_code = r.get(code_col)
                    try:
                        key = int(float(str(raw_code)))
                    except Exception:
                        key = str(raw_code).strip()
                    code_to_vehicle_name[key] = str(r.get(name_col, '')).strip()
    except Exception:
        code_to_vehicle_name = {}

    dados_volume = []

    # Group by combination of Agrupamento, Data Coleta, Veículo and Planta.
    # TIPO SATURACAO is not a grouping key to allow aggregation of different types into a single row.
    group_cols = ['AGRUPAMENTO', 'DATA COLETA','VEÍCULO']
    # Only add veiculo_col if it's distinct from existing group_cols

    # We fillna and strip to ensure groupby works properly on these columns
    for col in group_cols:
        if template[col].dtype == 'object':
            template[col] = template[col].str.strip()
        template[col] = template[col].fillna('')

    grouped = template.groupby(group_cols)
    
    
    for name, group in grouped:
        # name is a tuple matching group_cols
        group_dict = dict(zip(group_cols, name))
        
        volume_total = group['M³'].sum() if 'M³' in group.columns else 0
        peso_total = group['PESO TOTAL'].sum() if 'PESO TOTAL' in group.columns else 0
        embalagens_total = group['QTD EMBALAGENS'].sum() if 'QTD EMBALAGENS' in group.columns else 0
        
        # Determine saturation type for this aggregated group
        use_volume_saturation_for_group = True # Default
        tipo_saturacao_for_output = 'VOLUME'

        # Priority 1: Check for LINEHAUL rule (if multiple rows and FLW is LINEHAUL, force VOLUME)
        if 'FLW' in group.columns and len(group) > 1:
            print("Determine type for line haul or more")
            if group['FLW'].astype(str).str.strip().str.upper().eq('LINEHAUL').any():
                use_volume_saturation_for_group = True
                tipo_saturacao_for_output = 'VOLUME'
            else:
                # If multiple rows, but not LINEHAUL: "other remais as the they are"
                # Determine based on the most frequent TIPO SATURACAO in the group
                if 'TIPO SATURACAO' in group.columns and not group['TIPO SATURACAO'].dropna().empty:
                    most_common_tipo = group['TIPO SATURACAO'].value_counts().idxmax()
                    if 'PESO' in str(most_common_tipo).upper().strip():
                        use_volume_saturation_for_group = False
                        tipo_saturacao_for_output = 'PESO'
                    else: # Default to VOLUME for other cases (e.g., if most_common_tipo is 'VOLUME' or unrecognized)
                        use_volume_saturation_for_group = True
                        tipo_saturacao_for_output = 'VOLUME'
                # If TIPO SATURACAO column is not present or all NaN/empty, default to VOLUME
                else:
                    use_volume_saturation_for_group = True
                    tipo_saturacao_for_output = 'VOLUME'
        else: # Group has 1 row or 'FLW' column not present: "other remais as the they are"
            # Determine based on the TIPO SATURACAO of the single row or the first row if FLW is not relevant
            if 'TIPO SATURACAO' in group.columns and not group['TIPO SATURACAO'].dropna().empty:
                single_row_tipo = group['TIPO SATURACAO'].dropna().iloc[0] # Take the first non-null
                if 'PESO' in str(single_row_tipo).upper().strip():
                    use_volume_saturation_for_group = False
                    tipo_saturacao_for_output = 'PESO'
                else:
                    use_volume_saturation_for_group = True
                    tipo_saturacao_for_output = 'VOLUME'
            # Default to VOLUME if no TIPO SATURACAO column or no valid value.
            else:
                use_volume_saturation_for_group = True
                tipo_saturacao_for_output = 'VOLUME'

        if use_volume_saturation_for_group:
            saturacao_total = group['SAT VOLUME (%)'].sum() if 'SAT VOLUME (%)' in group.columns else 0
            coluna_sat = 'SAT VOLUME (%)'
        else:
            saturacao_total = group['SAT PESO (%)'].sum() if 'SAT PESO (%)' in group.columns else 0
            coluna_sat = 'SAT PESO (%)'

        # Calculate Apuração MDR
        total_desenhos = group['DESENHO'].nunique() if 'DESENHO' in group.columns else 0
        if coluna_sat in group.columns:
            desenhos_apurados = group[group[coluna_sat].fillna(0) > 0]['DESENHO'].nunique()
        else:
            desenhos_apurados = 0
            
        perc_mdr = round((desenhos_apurados / total_desenhos) * 100, 1) if total_desenhos else 0.0
        
        # Vehicle resolution
        veiculo_base = group_dict.get(veiculo_col, '')
        veiculo_final = manual_veiculo if use_manual and manual_veiculo is not None else veiculo_base
        veiculo_display = veiculo_final
        try:
            if str(veiculo_final).strip() != '':
                code_int = int(float(veiculo_final))
                veiculo_display = code_to_vehicle_name.get(code_int, veiculo_final)
        except Exception:
            veiculo_display = veiculo_final

        
        # Get one reference COD IMS, COD FORNECEDOR, and FORNECEDOR from the group
        # Prioritize non-null values
        ref_cod_ims = group['COD IMS'].dropna().iloc[0] if not group['COD IMS'].dropna().empty else ''
        ref_cod_fornecedor = group['COD FORNECEDOR'].dropna().iloc[0] if not group['COD FORNECEDOR'].dropna().empty else ''
        ref_fornecedor_name = group['FORNECEDOR'].dropna().iloc[0] if not group['FORNECEDOR'].dropna().empty else ''
        ref_plant = group['PLANTA'].dropna().iloc[0] if not group['PLANTA'].dropna().empty else ''

        dados_volume.append({
            'AGRUPAMENTO': group_dict.get('AGRUPAMENTO', ''),
            'DATA COLETA': group_dict.get('DATA COLETA', ''),
            'COD IMS': ref_cod_ims,
            'COD FORNECEDOR': ref_cod_fornecedor,
            'FORNECEDOR': ref_fornecedor_name,
            'VEÍCULO': veiculo_display,
            'TIPO DE SATURAÇÃO': tipo_saturacao_for_output, # Changed here
            'SATURAÇÃO TOTAL': round(saturacao_total, 2),
            'M³': round(volume_total, 3),
            'PESO': round(peso_total, 1),
            'Embalagens': int(embalagens_total),
            'APURAÇÃO MDR': perc_mdr,
            'PLANTA': ref_plant
        })

    df_volume = pd.DataFrame(dados_volume)
    df_volume = df_volume[df_volume[ 'SATURAÇÃO TOTAL' ] > 0]
    df_volume_zeros = df_volume[df_volume[ 'SATURAÇÃO TOTAL' ] == 0 ]
    
# --- Save with blue header formatting and auto-width columns ---
    with pd.ExcelWriter('Volume_por_rota.xlsx', engine='openpyxl') as writer:

        df_volume.to_excel(writer, sheet_name='Volume por Rota', index=False)
        df_volume_zeros.to_excel(writer, sheet_name='Saturação Nulos', index=False)

        header_fill = PatternFill(
            start_color='00246C',
            end_color='00246C',
            fill_type='solid'
        )
        header_font = Font(bold=True, color='FFFFFF')
        header_align = Alignment(horizontal='center', vertical='center')

        # Apply formatting to both sheets
        for sheet_name in ['Volume por Rota', 'Saturação Nulos']:

            ws = writer.sheets[sheet_name]

            for col_num, col in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):

                max_length = 0
                column_letter = get_column_letter(col_num)

                header_cell = ws[f'{column_letter}1']

                if header_cell.value:
                    max_length = len(str(header_cell.value))

                for row_num in range(2, min(102, ws.max_row + 1)):
                    cell = ws[f'{column_letter}{row_num}']

                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

                header_cell.fill = header_fill
                header_cell.font = header_font
                header_cell.alignment = header_align
    #tree = ttk.Treeview()
    #tree_resumo = ttk.Treeview()
#completar_informacoes(tree,3, tree_resumo)